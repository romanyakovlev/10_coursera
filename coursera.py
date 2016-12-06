from lxml import etree
from bs4 import BeautifulSoup
from datetime import date
from openpyxl import Workbook
import requests
import random
import json


def get_courses_list():
    courses_amount = 20
    xml_page = requests.get('https://www.coursera.org/sitemap~www~courses.xml')
    xml_root = etree.fromstring(xml_page.content)
    random_courses = [course[0].text for course in random.sample(set(xml_root), courses_amount)]
    return random_courses


def get_course_info(course_slug):
    full_path = 'https://www.coursera.org/learn/{}'.format(course_slug)
    html_doc = requests.get(full_path).content
    soup_data = BeautifulSoup(html_doc, 'html.parser')
    request_params = {"q": "slug", "slug": course_slug,
                      "fields": "plannedLaunchDate,upcomingSessionStartDate"}
    response_json = json.loads(requests.get("https://api.coursera.org/api/courses.v1/",
                                            params=request_params).text)
    course_info = response_json['elements'][0]
    if 'upcomingSessionStartDate' in course_info.keys():
        timestamp_in_millisecs = course_info['upcomingSessionStartDate']
        millisecs_in_seconds = 1000.0
        timestamp_in_secs = timestamp_in_millisecs / millisecs_in_seconds
        course_start_date = date.fromtimestamp(timestamp_in_secs)
    else:
        course_start_date = course_info['plannedLaunchDate']
    rating_data = soup_data.find(class_='ratings-text')
    if rating_data:
        course_rating = soup_data.find(class_='ratings-text').text
    else:
        course_rating = 'No rating'
    course_name = soup_data.find(class_='course-name-text').text
    course_language = soup_data.find(class_='language-info').text
    weeks_list = soup_data.find_all(class_='week')
    course_weeks_amount = len(weeks_list) if weeks_list else 'No information'
    return course_name, course_rating, course_language, course_start_date, course_weeks_amount


def output_courses_info_to_xlsx(filepath):
    workbook = Workbook()
    worksheet = workbook.active
    fields_name = iter(['Name', 'Rating', 'Language', 'Start Date','Weeks Amount'])
    min_row_num, max_row_num, max_col_num  = 1, 1, 5
    for cells in worksheet.iter_rows(max_row=max_row_num,
                                     max_col=max_col_num, min_row=min_row_num):
        for cell in cells:
            cell.value = next(fields_name)
    coursers_list = get_courses_list()
    second_row, last_row = 2, 21
    for cells, course in zip(
            worksheet.iter_rows(max_row=last_row, max_col=max_col_num, min_row=second_row),
            coursers_list):
        course_slug = course.split('/')[-1]
        course_params = get_course_info(course_slug)
        for cell, parameter in zip(cells, course_params):
            cell.value = parameter
    workbook.save(filepath)


if __name__ == '__main__':
    filepath = input('Enter file that will consist information about courses:\n')
    print('Start parsing of 20 any random Coursera courses. You need to wait a little bit...')
    output_courses_info_to_xlsx(filepath)
    print('Done. Data stored to {}'.format(filepath))
